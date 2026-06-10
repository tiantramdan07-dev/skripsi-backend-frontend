# ============================================================
# Timbangan Digital AI — Flask Backend (Production v3.0)
# PT Interskala Mandiri Indonesia
# IP Server: 10.183.165.82 | Port: 4000
# ============================================================
import os, time, uuid, base64, random, decimal, threading, io
import numpy as np
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError
from threading import Lock
from functools import wraps

import cv2
from ultralytics import YOLO
import psycopg2
import psycopg2.extras
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from flask import Flask, jsonify, request, send_from_directory, send_file
from flask_cors import CORS
import jwt

# ─── CONFIG ──────────────────────────────────────────────────
DB_CONFIG = {
    "host":     os.environ.get("DB_HOST",   "localhost"),
    "port":     int(os.environ.get("DB_PORT", 5432)),
    "user":     os.environ.get("DB_USER",   "postgres"),
    "password": os.environ.get("DB_PASS",   "gajahbengkak"),
    "dbname":   os.environ.get("DB_NAME",   "timbangandigital_ai"),
}
SECRET_KEY     = os.environ.get("SECRET_KEY",  "timbangandigital-produksi-2025-interskala")
JWT_ALGO       = "HS256"
JWT_EXP_HOURS  = int(os.environ.get("JWT_EXP_HOURS", 8))
SIMULATE_SCALE = os.environ.get("SIMULATE_SCALE", "0") == "1"
SERIAL_PORT    = os.environ.get("SERIAL_PORT", "COM8")
BAUD_RATE      = int(os.environ.get("BAUD_RATE", 9600))

UPLOAD_FOLDER = os.path.join(os.getcwd(), "static", "assets", "img")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
ALLOWED_EXT = {"png", "jpg", "jpeg", "gif", "webp"}

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}}, supports_credentials=True)

# ─── DB HELPERS ──────────────────────────────────────────────
def get_db():
    return psycopg2.connect(**DB_CONFIG)

def allowed_file(fname):
    return "." in fname and fname.rsplit(".", 1)[1].lower() in ALLOWED_EXT

def as_number(x):
    try:   return float(x.cpu().numpy())
    except:
        try: return float(x)
        except: return 0.0

def row_to_dict(row):
    d = dict(row)
    for k, v in d.items():
        if isinstance(v, decimal.Decimal): d[k] = float(v)
        elif isinstance(v, datetime):      d[k] = v.isoformat()
    return d

def wib_now():
    return datetime.now(tz=ZoneInfo("Asia/Jakarta"))

# ─── DB SCHEMA AUTO-MIGRATE ──────────────────────────────────
def ensure_schema():
    """Create tables if not exist & ensure users.is_active column exists."""
    db = get_db(); cur = db.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id            SERIAL PRIMARY KEY,
            first_name    VARCHAR(100) DEFAULT '',
            last_name     VARCHAR(100) DEFAULT '',
            email         VARCHAR(255) UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role          VARCHAR(20)  DEFAULT 'operator',
            is_active     BOOLEAN      DEFAULT TRUE,
            created_at    TIMESTAMPTZ  DEFAULT NOW()
        )
    """)
    # Add is_active if missing (migration)
    cur.execute("""
        ALTER TABLE users ADD COLUMN IF NOT EXISTS is_active BOOLEAN DEFAULT TRUE
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS produk (
            kode_produk SERIAL PRIMARY KEY,
            nama_produk VARCHAR(200) NOT NULL,
            harga_per_kg NUMERIC(12,2) NOT NULL,
            path_gambar TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS transaksi (
            id          SERIAL PRIMARY KEY,
            nama_produk VARCHAR(200),
            berat_kg    NUMERIC(10,3),
            harga_per_kg NUMERIC(12,2),
            total_harga  NUMERIC(14,2),
            timestamp   TIMESTAMPTZ DEFAULT NOW()
        )
    """)
    db.commit(); cur.close(); db.close()
    print("✅ Schema ready")

# ─── YOLO MODEL ──────────────────────────────────────────────
MODEL_PATH = os.environ.get("MODEL_PATH", os.path.join(os.getcwd(), "models", "best.pt"))
try:
    model = YOLO(MODEL_PATH)
    print(f"✅ Model YOLO loaded: {MODEL_PATH}")
except Exception as e:
    print(f"⚠️  Gagal load model YOLO: {e}")
    model = None

# ─── GLOBAL STATE ────────────────────────────────────────────
latest_weight      = 0.0
latest_detection   = {}
latest_esp32_frame = None
weight_lock        = Lock()
client_lock        = Lock()
esp32_frame_lock   = Lock()
client_last_ts     = {}
MIN_INTERVAL_S     = 0.06
ESP32_CLIENT_ID    = "esp32-cam-01"
WORKER_POOL        = ThreadPoolExecutor(max_workers=4)

# ─── JWT & AUTH DECORATORS ───────────────────────────────────
def create_token(user):
    payload = {
        "id":    user["id"],
        "email": user["email"],
        "role":  user["role"],
        "exp":   datetime.utcnow() + timedelta(hours=JWT_EXP_HOURS),
        "iat":   datetime.utcnow(),
    }
    return jwt.encode(payload, SECRET_KEY, algorithm=JWT_ALGO)

def _decode_token():
    token = None
    if "Authorization" in request.headers:
        parts = request.headers["Authorization"].split()
        if len(parts) == 2: token = parts[1]
    if not token:
        token = request.args.get("token")
    if not token:
        return None, None
    try:
        data = jwt.decode(token, SECRET_KEY, algorithms=[JWT_ALGO])
        return data["email"], data.get("role", "operator")
    except:
        return None, None

def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        email, role = _decode_token()
        if not email:
            return jsonify({"message": "Token missing or invalid!"}), 401
        return f(email, *args, **kwargs)
    return decorated

def admin_required(f):
    """Only users with role='admin' can access."""
    @wraps(f)
    def decorated(*args, **kwargs):
        email, role = _decode_token()
        if not email:
            return jsonify({"message": "Token missing or invalid!"}), 401
        if role != "admin":
            return jsonify({"message": "Akses ditolak: hanya Admin yang diperbolehkan"}), 403
        return f(email, *args, **kwargs)
    return decorated

# ─── SERIAL / SCALE THREAD ───────────────────────────────────
def read_scale_serial():
    global latest_weight
    if SIMULATE_SCALE:
        print("⚠️  Scale: SIMULASI aktif")
        while True:
            with weight_lock: latest_weight = round(random.uniform(0.05, 3.0), 3)
            time.sleep(0.5)
        return
    try:
        import serial
        ser = serial.Serial(SERIAL_PORT, BAUD_RATE, timeout=1)
        time.sleep(2)
        print(f"✅ Timbangan serial terhubung di {SERIAL_PORT}")
        while True:
            try:
                line = ser.readline().decode("utf-8").strip()
                if not line: continue
                w = float(line)
                if w < 0: w = 0.0
                with weight_lock: latest_weight = round(w, 3)
            except ValueError: pass
            except Exception as e: print("❌ Serial error:", e)
            time.sleep(0.1)
    except Exception as e:
        print(f"❌ Gagal konek serial: {e}")

# ─── YOLO INFERENCE ──────────────────────────────────────────
def process_frame_yolo(frame):
    if model is None:
        _, buf = cv2.imencode(".jpg", frame)
        return "Model not loaded", [], buf.tobytes()
    results   = model(frame, conf=0.6, verbose=False)
    dets      = []
    top_label = "Tidak ada"
    try: boxes = results[0].boxes
    except: boxes = []
    if len(boxes) > 0:
        for b in boxes:
            cls_idx = int(as_number(b.cls))
            conf    = float(as_number(b.conf))
            xyxy    = b.xyxy[0]
            x1,y1,x2,y2 = (float(as_number(xyxy[i])) for i in range(4))
            label   = model.names.get(cls_idx, str(cls_idx)) if hasattr(model, "names") else str(cls_idx)
            dets.append({
                "cls": cls_idx, "label": label,
                "conf": int(round(conf*100)), "confidence": int(round(conf*100)),
                "x": round(x1,1), "y": round(y1,1),
                "width": round(x2-x1,1), "height": round(y2-y1,1),
            })
        top_label = dets[0]["label"]
    try:    annotated = results[0].plot()
    except: annotated = frame
    _, buf = cv2.imencode(".jpg", annotated)
    return top_label, dets, buf.tobytes()

# ═══════════════════════════════════════════════════════════════
# AUTH ROUTES
# ═══════════════════════════════════════════════════════════════
@app.route("/")
def index():
    return jsonify({"status": "Timbangan Digital AI API", "version": "3.0", "server": "10.183.165.82"})

@app.route("/auth/signup", methods=["POST"])
def auth_signup():
    data  = request.get_json(force=True) or {}
    first = data.get("first_name","").strip()
    last  = data.get("last_name","").strip()
    email = (data.get("email") or "").strip().lower()
    pw    = data.get("password","")
    if not email or not pw:
        return jsonify({"error": "Email dan password wajib diisi"}), 400
    db  = get_db()
    cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id FROM users WHERE email=%s", (email,))
    if cur.fetchone():
        cur.close(); db.close()
        return jsonify({"error": "Email sudah terdaftar"}), 400
    pw_hash = generate_password_hash(pw)
    # First user ever → admin, else → operator
    cur.execute("SELECT COUNT(*) as c FROM users")
    count = cur.fetchone()["c"]
    role  = "admin" if count == 0 else "operator"
    cur.execute(
        "INSERT INTO users(first_name,last_name,email,password_hash,role) VALUES(%s,%s,%s,%s,%s) RETURNING id",
        (first, last, email, pw_hash, role)
    )
    user = cur.fetchone(); db.commit(); cur.close(); db.close()
    return jsonify({"message": "registered", "id": user["id"], "role": role}), 201

@app.route("/auth/login", methods=["POST"])
def auth_login():
    data  = request.get_json(force=True) or {}
    email = (data.get("email") or "").strip().lower()
    pw    = data.get("password","")
    if not email or not pw:
        return jsonify({"error": "Email dan password wajib diisi"}), 400
    db  = get_db()
    cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id,email,password_hash,role,first_name,last_name,is_active FROM users WHERE email=%s", (email,))
    u = cur.fetchone(); cur.close(); db.close()
    if not u or not check_password_hash(u["password_hash"], pw):
        return jsonify({"error": "Email atau password salah"}), 401
    if not u.get("is_active", True):
        return jsonify({"error": "Akun dinonaktifkan. Hubungi administrator."}), 403
    token = create_token(u)
    return jsonify({
        "token":      token,
        "email":      u["email"],
        "role":       u["role"],
        "first_name": u["first_name"],
        "last_name":  u["last_name"],
    })

@app.route("/auth/me", methods=["GET"])
@token_required
def auth_me(current_email):
    db  = get_db()
    cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id,first_name,last_name,email,role,is_active,created_at FROM users WHERE email=%s", (current_email,))
    u = cur.fetchone(); cur.close(); db.close()
    if not u: return jsonify({"error": "not_found"}), 404
    return jsonify(row_to_dict(u))

# ═══════════════════════════════════════════════════════════════
# USER MANAGEMENT (Admin Only)
# ═══════════════════════════════════════════════════════════════
@app.route("/api/users", methods=["GET"])
@admin_required
def api_users_list(current_email):
    db  = get_db()
    cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT id, first_name, last_name, email, role, is_active, created_at
        FROM users ORDER BY id ASC
    """)
    rows = [row_to_dict(r) for r in cur.fetchall()]
    cur.close(); db.close()
    return jsonify(rows)

@app.route("/api/users", methods=["POST"])
@admin_required
def api_users_create(current_email):
    data  = request.get_json(force=True) or {}
    first = data.get("first_name","").strip()
    last  = data.get("last_name","").strip()
    email = (data.get("email") or "").strip().lower()
    pw    = data.get("password","")
    role  = data.get("role","operator")
    if not email or not pw:
        return jsonify({"error": "Email dan password wajib"}), 400
    if role not in ("admin","operator"):
        return jsonify({"error": "Role tidak valid"}), 400
    db  = get_db()
    cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id FROM users WHERE email=%s", (email,))
    if cur.fetchone():
        cur.close(); db.close()
        return jsonify({"error": "Email sudah terdaftar"}), 400
    pw_hash = generate_password_hash(pw)
    cur.execute(
        "INSERT INTO users(first_name,last_name,email,password_hash,role,is_active) VALUES(%s,%s,%s,%s,%s,TRUE) RETURNING id",
        (first, last, email, pw_hash, role)
    )
    new = cur.fetchone(); db.commit(); cur.close(); db.close()
    return jsonify({"message": "User berhasil dibuat", "id": new["id"]}), 201

@app.route("/api/users/<int:user_id>", methods=["PUT"])
@admin_required
def api_users_update(current_email, user_id):
    data = request.get_json(force=True) or {}
    updates, values = [], []
    if "role" in data:
        if data["role"] not in ("admin","operator"):
            return jsonify({"error": "Role tidak valid"}), 400
        updates.append("role=%s"); values.append(data["role"])
    if "is_active" in data:
        updates.append("is_active=%s"); values.append(bool(data["is_active"]))
    if "first_name" in data:
        updates.append("first_name=%s"); values.append(data["first_name"].strip())
    if "last_name" in data:
        updates.append("last_name=%s"); values.append(data["last_name"].strip())
    if not updates:
        return jsonify({"error": "Tidak ada data untuk diupdate"}), 400

    # Prevent admin from demoting themselves
    db  = get_db()
    cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT email FROM users WHERE id=%s", (user_id,))
    target = cur.fetchone()
    if not target:
        cur.close(); db.close()
        return jsonify({"error": "User tidak ditemukan"}), 404
    if target["email"] == current_email and "role" in data and data["role"] != "admin":
        cur.close(); db.close()
        return jsonify({"error": "Tidak bisa menurunkan role diri sendiri"}), 400

    values.append(user_id)
    cur.execute(f"UPDATE users SET {','.join(updates)} WHERE id=%s RETURNING id,first_name,last_name,email,role,is_active",
                tuple(values))
    updated = row_to_dict(cur.fetchone()); db.commit(); cur.close(); db.close()
    return jsonify(updated)

@app.route("/api/users/<int:user_id>", methods=["DELETE"])
@admin_required
def api_users_delete(current_email, user_id):
    db  = get_db()
    cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT email FROM users WHERE id=%s", (user_id,))
    target = cur.fetchone()
    if not target:
        cur.close(); db.close()
        return jsonify({"error": "User tidak ditemukan"}), 404
    if target["email"] == current_email:
        cur.close(); db.close()
        return jsonify({"error": "Tidak bisa menghapus akun sendiri"}), 400
    cur.execute("DELETE FROM users WHERE id=%s", (user_id,))
    db.commit(); cur.close(); db.close()
    return jsonify({"message": f"User {user_id} berhasil dihapus"})

@app.route("/api/users/<int:user_id>/reset-password", methods=["POST"])
@admin_required
def api_users_reset_password(current_email, user_id):
    data = request.get_json(force=True) or {}
    new_pw = data.get("password","").strip()
    if len(new_pw) < 6:
        return jsonify({"error": "Password minimal 6 karakter"}), 400
    db  = get_db()
    cur = db.cursor()
    cur.execute("UPDATE users SET password_hash=%s WHERE id=%s RETURNING id",
                (generate_password_hash(new_pw), user_id))
    if not cur.fetchone():
        cur.close(); db.close()
        return jsonify({"error": "User tidak ditemukan"}), 404
    db.commit(); cur.close(); db.close()
    return jsonify({"message": "Password berhasil direset"})

# ═══════════════════════════════════════════════════════════════
# ESP32 WEIGHT PUSH
# ═══════════════════════════════════════════════════════════════
@app.route("/api/weight", methods=["POST"])
def api_weight():
    global latest_weight
    data = request.get_json(silent=True) or {}
    try:
        w = float(data.get("weight", 0))
        if w < 0: w = 0.0
        with weight_lock: latest_weight = round(w, 3)
        return jsonify({"ok": True, "weight": latest_weight})
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route("/api/weight", methods=["GET"])
def api_weight_get():
    with weight_lock: w = latest_weight
    return jsonify({"weight": w, "ts": wib_now().isoformat()})

# ═══════════════════════════════════════════════════════════════
# ESP32-S3 CAM FRAME PUSH (Device endpoint — no JWT)
# ═══════════════════════════════════════════════════════════════
@app.route("/api/esp32_frame", methods=["POST"])
def api_esp32_frame():
    global latest_detection, latest_esp32_frame
    now_ts = time.time()
    with client_lock:
        if now_ts - client_last_ts.get(ESP32_CLIENT_ID, 0) < MIN_INTERVAL_S:
            return jsonify({"ok": False, "reason": "rate_limited"}), 429
        client_last_ts[ESP32_CLIENT_ID] = now_ts

    frame = None
    ct = request.content_type or ""
    if "image/jpeg" in ct or "application/octet-stream" in ct:
        nparr = np.frombuffer(request.data, np.uint8)
        frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    elif "multipart/form-data" in ct:
        f = request.files.get("frame")
        if f:
            nparr = np.frombuffer(f.read(), np.uint8)
            frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    else:
        data = request.get_json(force=True, silent=True) or {}
        b64  = data.get("frame","")
        if not b64: return jsonify({"error": "no_frame"}), 400
        try:
            if "," in b64: b64 = b64.split(",",1)[1]
            nparr = np.frombuffer(base64.b64decode(b64), np.uint8)
            frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        except Exception as e:
            return jsonify({"error": "decode_failed", "detail": str(e)}), 400

    if frame is None:
        return jsonify({"error": "cannot_decode_frame"}), 400

    fut = WORKER_POOL.submit(process_frame_yolo, frame)
    try:
        label, boxes, ann_bytes = fut.result(timeout=10)
    except FuturesTimeoutError:
        return jsonify({"error": "yolo_timeout"}), 504
    except Exception as e:
        return jsonify({"error": "yolo_error", "detail": str(e)}), 500

    with weight_lock: w = latest_weight
    ann_b64 = "data:image/jpeg;base64," + base64.b64encode(ann_bytes).decode()

    with esp32_frame_lock: latest_esp32_frame = ann_b64
    latest_detection[ESP32_CLIENT_ID] = {
        "detection":       label,
        "weight":          w,
        "ts":              wib_now().isoformat(),
        "annotated_frame": ann_b64,
    }
    return jsonify({"ok": True, "detection": label, "weight": w, "boxes": len(boxes)})

# ═══════════════════════════════════════════════════════════════
# STATUS (polling by frontend)
# ═══════════════════════════════════════════════════════════════
@app.route("/api/status", methods=["POST"])
def api_status():
    global latest_detection, latest_weight, latest_esp32_frame
    data      = request.get_json(silent=True) or {}
    client_id = data.get("client_id", ESP32_CLIENT_ID)
    cached    = latest_detection.get(ESP32_CLIENT_ID) or latest_detection.get(client_id)
    with weight_lock:   w   = latest_weight
    with esp32_frame_lock: ann = latest_esp32_frame
    if not cached:
        return jsonify({"detection": "-", "weight": w, "annotated_frame": ann, "ts": wib_now().isoformat()})
    return jsonify({
        "detection":       cached.get("detection", "-"),
        "weight":          w,
        "annotated_frame": ann,
        "ts":              wib_now().isoformat(),
    })

# ═══════════════════════════════════════════════════════════════
# DETECT FRAME (browser fallback)
# ═══════════════════════════════════════════════════════════════
@app.route("/api/detect_frame", methods=["POST"])
@token_required
def api_detect_frame(current_email):
    global latest_detection, latest_weight
    data = request.get_json(force=True, silent=True)
    if not data: return jsonify({"error": "invalid_json"}), 400
    frame_b64 = data.get("frame") or data.get("image")
    client_id = data.get("client_id") or str(uuid.uuid4())
    if not frame_b64: return jsonify({"error": "no_frame"}), 400
    now_ts = time.time()
    with client_lock:
        if now_ts - client_last_ts.get(client_id, 0) < MIN_INTERVAL_S:
            return jsonify({"error": "too_many_requests"}), 429
        client_last_ts[client_id] = now_ts
    try:
        b64   = frame_b64.split(",",1)[1] if "," in frame_b64 else frame_b64
        nparr = np.frombuffer(base64.b64decode(b64), np.uint8)
        frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        if frame is None: raise ValueError("cannot decode")
    except Exception as e:
        return jsonify({"error": "decode_failed", "detail": str(e)}), 400
    fut = WORKER_POOL.submit(process_frame_yolo, frame)
    try:
        label, boxes, ann_bytes = fut.result(timeout=12)
    except FuturesTimeoutError:
        return jsonify({"error": "timeout"}), 504
    except Exception as e:
        return jsonify({"error": "yolo_error", "detail": str(e)}), 500
    with weight_lock: w = latest_weight
    ann_b64 = "data:image/jpeg;base64," + base64.b64encode(ann_bytes).decode()
    latest_detection[client_id] = {"detection": label, "weight": w, "ts": wib_now().isoformat(), "annotated_frame": ann_b64}
    return jsonify({"detection": label, "boxes": boxes, "annotated_frame": ann_b64, "weight": w, "server_time": wib_now().isoformat()})

# ═══════════════════════════════════════════════════════════════
# PRODUK CRUD
# ═══════════════════════════════════════════════════════════════
@app.route("/api/produk", methods=["GET"])
def api_get_produk():
    db  = get_db()
    cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT kode_produk,nama_produk,harga_per_kg,path_gambar FROM produk ORDER BY kode_produk")
    rows = [row_to_dict(r) for r in cur.fetchall()]
    cur.close(); db.close()
    return jsonify(rows)

@app.route("/api/produk/<int:kode>", methods=["GET"])
def api_get_produk_one(kode):
    db  = get_db()
    cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM produk WHERE kode_produk=%s", (kode,))
    r = cur.fetchone(); cur.close(); db.close()
    if not r: return jsonify({"error": "not found"}), 404
    return jsonify(row_to_dict(r))

@app.route("/api/produk", methods=["POST"])
@admin_required
def api_create_produk(current_email):
    is_form = request.content_type and "multipart/form-data" in request.content_type
    body    = {} if is_form else (request.get_json(force=True) or {})
    nama    = request.form.get("nama_produk") if is_form else body.get("nama_produk")
    harga   = request.form.get("harga_per_kg") if is_form else body.get("harga_per_kg")
    file    = request.files.get("gambar") if is_form else None
    if not nama or harga is None: return jsonify({"error": "nama_produk dan harga_per_kg wajib"}), 400
    try:
        harga_f = float(harga)
        if harga_f < 0: raise ValueError
    except: return jsonify({"error": "harga_per_kg harus angka positif"}), 400
    path_gambar = None
    if file and file.filename:
        if not allowed_file(file.filename): return jsonify({"error": "Format file tidak diizinkan"}), 400
        nm, ext  = os.path.splitext(secure_filename(file.filename))
        filename = f"{nm}_{datetime.now().strftime('%Y%m%d%H%M%S%f')}{ext}"
        file.save(os.path.join(UPLOAD_FOLDER, filename))
        path_gambar = f"/static/assets/img/{filename}"
    db  = get_db(); cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("INSERT INTO produk(nama_produk,harga_per_kg,path_gambar) VALUES(%s,%s,%s) RETURNING *",
                (nama, harga_f, path_gambar))
    new = row_to_dict(cur.fetchone()); db.commit(); cur.close(); db.close()
    return jsonify(new), 201

@app.route("/api/produk/<int:kode>", methods=["PUT"])
@admin_required
def api_update_produk(current_email, kode):
    is_form = request.content_type and "multipart/form-data" in request.content_type
    body    = {} if is_form else (request.get_json(force=True) or {})
    nama  = request.form.get("nama_produk") if is_form else body.get("nama_produk")
    harga = request.form.get("harga_per_kg") if is_form else body.get("harga_per_kg")
    file  = request.files.get("gambar") if is_form else None
    updates, values = [], []
    if nama:  updates.append("nama_produk=%s"); values.append(nama)
    if harga:
        try:
            hf = float(harga)
            if hf < 0: raise ValueError
        except: return jsonify({"error": "harga tidak valid"}), 400
        updates.append("harga_per_kg=%s"); values.append(hf)
    if file and file.filename:
        if not allowed_file(file.filename): return jsonify({"error": "format tidak diizinkan"}), 400
        nm, ext  = os.path.splitext(secure_filename(file.filename))
        filename = f"{nm}_{datetime.now().strftime('%Y%m%d%H%M%S%f')}{ext}"
        file.save(os.path.join(UPLOAD_FOLDER, filename))
        updates.append("path_gambar=%s"); values.append(f"/static/assets/img/{filename}")
    if not updates: return jsonify({"error": "tidak ada data untuk diupdate"}), 400
    values.append(kode)
    db  = get_db(); cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(f"UPDATE produk SET {','.join(updates)} WHERE kode_produk=%s RETURNING *", tuple(values))
    r = cur.fetchone(); db.commit(); cur.close(); db.close()
    if not r: return jsonify({"error": "not found"}), 404
    return jsonify(row_to_dict(r))

@app.route("/api/produk/<int:kode>", methods=["DELETE"])
@admin_required
def api_delete_produk(current_email, kode):
    db  = get_db(); cur = db.cursor()
    cur.execute("DELETE FROM produk WHERE kode_produk=%s RETURNING kode_produk", (kode,))
    d = cur.fetchone(); db.commit(); cur.close(); db.close()
    if not d: return jsonify({"error": "not found"}), 404
    return jsonify({"message": f"Produk {kode} dihapus"})

@app.route("/static/assets/img/<path:filename>")
def serve_image(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)

# ═══════════════════════════════════════════════════════════════
# TRANSAKSI / CETAK
# ═══════════════════════════════════════════════════════════════
@app.route("/cetak", methods=["POST"])
@token_required
def cetak(current_email):
    data     = request.get_json(force=True) or {}
    required = ["nama_produk","berat_kg","harga_per_kg","total_harga"]
    for f in required:
        if f not in data: return jsonify({"status": f"Field {f} missing"}), 400
    db  = get_db(); cur = db.cursor()
    cur.execute(
        "INSERT INTO transaksi(nama_produk,berat_kg,harga_per_kg,total_harga,timestamp) VALUES(%s,%s,%s,%s,%s) RETURNING id",
        (data["nama_produk"], data["berat_kg"], data["harga_per_kg"], data["total_harga"], wib_now())
    )
    trx_id = cur.fetchone()[0]; db.commit(); cur.close(); db.close()
    return jsonify({"status": f"Transaksi {data['nama_produk']} berhasil disimpan!", "trx_id": trx_id, "ok": True})

def build_escpos(nama_produk, berat_kg, harga_per_kg, total_harga, waktu, kasir="Kasir"):
    ESC = b'\x1b'; GS = b'\x1d'
    def center(text, width=32): return text.center(width)[:width]
    def line(width=32): return "-" * width
    buf = bytearray()
    buf += ESC + b'@'
    buf += ESC + b'!' + bytes([0x30])
    buf += (center("TIMBANGAN DIGITAL AI") + "\n").encode("ascii", errors="replace")
    buf += ESC + b'!' + bytes([0x00])
    buf += (center("PT INTERSKALA MANDIRI IND") + "\n").encode("ascii", errors="replace")
    buf += (line() + "\n").encode()
    rows = [
        ("PRODUK", nama_produk), ("BERAT ", f"{berat_kg:.3f} kg"),
        ("HARGA ", f"Rp {int(harga_per_kg):,}".replace(",",".")+"/kg"),
        ("TOTAL ", f"Rp {int(total_harga):,}".replace(",",".")),
        ("WAKTU ", waktu[:19]), ("KASIR ", kasir),
    ]
    for label, value in rows:
        buf += (f"{label}: {value}\n").encode("ascii", errors="replace")
    buf += (line() + "\n").encode()
    buf += (center("Terima Kasih!") + "\n").encode("ascii", errors="replace")
    buf += (center("PT Interskala Mandiri Indonesia") + "\n").encode("ascii", errors="replace")
    buf += b'\n\n\n\n\n\n'
    buf += ESC + b'd' + bytes([8])   # dari 4 → 8
    buf += GS + b'V' + bytes([0x41, 0x00])
    return bytes(buf)

# @app.route("/api/print_rawbt", methods=["POST"])
# @token_required
# def api_print_rawbt(current_email):
#     data = request.get_json(force=True) or {}
#     try:
#         nama  = str(data["nama_produk"])
#         berat = float(data["berat_kg"])
#         harga = int(data["harga_per_kg"])
#         total = int(data["total_harga"])
#     except (KeyError, ValueError) as e:
#         return jsonify({"error": f"invalid field: {e}"}), 400
#     db  = get_db(); cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
#     cur.execute("SELECT first_name,last_name FROM users WHERE email=%s", (current_email,))
#     u = cur.fetchone(); cur.close(); db.close()
#     kasir  = f"{u['first_name']} {u['last_name']}".strip() if u else current_email.split("@")[0]
#     waktu  = wib_now().strftime("%d/%m/%Y %H:%M:%S")
#     escpos = build_escpos(nama, berat, harga, total, waktu, kasir)
#     b64    = base64.b64encode(escpos).decode()
#     return jsonify({"rawbt_uri": f"rawbt://base64/{b64}", "escpos_b64": b64, "kasir": kasir, "waktu": waktu})

# ═══════════════════════════════════════════════════════════════
# RIWAYAT & LAPORAN
# ═══════════════════════════════════════════════════════════════
@app.route("/api/riwayat", methods=["GET"])
@token_required
def get_riwayat(current_email):
    tanggal = request.args.get("tanggal","")
    sort    = request.args.get("sort","desc")
    db  = get_db(); cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    q   = "SELECT id,nama_produk,berat_kg AS berat,harga_per_kg,total_harga,timestamp AS waktu FROM transaksi"
    p   = []
    if tanggal:
        if ":" in tanggal:
            s,e = tanggal.split(":",1)
            q += " WHERE DATE(timestamp) BETWEEN %s AND %s"; p += [s,e]
        else:
            q += " WHERE DATE(timestamp)=%s"; p.append(tanggal)
    q += f" ORDER BY timestamp {'ASC' if sort=='asc' else 'DESC'}"
    cur.execute(q, tuple(p)); rows = [row_to_dict(r) for r in cur.fetchall()]
    cur.close(); db.close()
    return jsonify(rows)

@app.route("/api/laporan/export/excel", methods=["GET"])
@token_required
def export_excel(current_email):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    tanggal = request.args.get("tanggal","")
    db  = get_db(); cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    q   = "SELECT id,nama_produk,berat_kg,harga_per_kg,total_harga,timestamp FROM transaksi"
    p   = []
    if tanggal:
        if ":" in tanggal:
            s,e = tanggal.split(":",1); q += " WHERE DATE(timestamp) BETWEEN %s AND %s"; p+=[s,e]
        else: q += " WHERE DATE(timestamp)=%s"; p.append(tanggal)
    q += " ORDER BY timestamp DESC"
    cur.execute(q, tuple(p)); rows = cur.fetchall(); cur.close(); db.close()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Laporan Penimbangan"
    hf = PatternFill(fill_type="solid", fgColor="1F4E79"); hfont = Font(bold=True, color="FFFFFF")
    ws.append(["No","Nama Produk","Berat (Kg)","Harga/Kg (Rp)","Total Harga (Rp)","Waktu"])
    for cell in ws[1]: cell.fill = hf; cell.font = hfont; cell.alignment = Alignment(horizontal="center")
    for i, r in enumerate(rows, 1):
        ts = r["timestamp"].strftime("%d/%m/%Y %H:%M") if isinstance(r["timestamp"], datetime) else str(r["timestamp"])
        ws.append([i, r["nama_produk"], float(r["berat_kg"]), float(r["harga_per_kg"]), float(r["total_harga"]), ts])
    ws.column_dimensions["B"].width = 25; ws.column_dimensions["F"].width = 22
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"laporan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/laporan/export/pdf", methods=["GET"])
@token_required
def export_pdf(current_email):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    tanggal = request.args.get("tanggal","")
    db  = get_db(); cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    q   = "SELECT id,nama_produk,berat_kg,harga_per_kg,total_harga,timestamp FROM transaksi"
    p   = []
    if tanggal:
        if ":" in tanggal:
            s,e = tanggal.split(":",1); q += " WHERE DATE(timestamp) BETWEEN %s AND %s"; p+=[s,e]
        else: q += " WHERE DATE(timestamp)=%s"; p.append(tanggal)
    q += " ORDER BY timestamp DESC"
    cur.execute(q, tuple(p)); rows = cur.fetchall(); cur.close(); db.close()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4); styles = getSampleStyleSheet(); elems = []
    elems.append(Paragraph("Laporan Penimbangan — PT Interskala Mandiri Indonesia", styles["Title"]))
    elems.append(Spacer(1,12))
    tdata = [["No","Nama Produk","Berat (Kg)","Harga/Kg","Total Harga","Waktu"]]
    for i, r in enumerate(rows, 1):
        ts = r["timestamp"].strftime("%d/%m/%Y %H:%M") if isinstance(r["timestamp"], datetime) else str(r["timestamp"])
        tdata.append([i, r["nama_produk"], f"{float(r['berat_kg']):.3f}",
                      f"Rp {int(r['harga_per_kg']):,}".replace(",","."),
                      f"Rp {int(r['total_harga']):,}".replace(",","."), ts])
    t = Table(tdata, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1F4E79")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#EFF3F8")]),
        ("GRID",(0,0),(-1,-1),0.5,colors.grey),
    ]))
    elems.append(t); doc.build(elems); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"laporan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                     mimetype="application/pdf")

# ═══════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    ensure_schema()
    t = threading.Thread(target=read_scale_serial, daemon=True)
    t.start()
    print("🚀 Server running on http://10.183.165.82:4000")
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 4000)),
            debug=False, use_reloader=False, threaded=True)
